import rrdtool
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook 
import time
from pprint import pprint
import re
import smtplib
from string import Template
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase
from email import encoders
import sys

def average(lst):
   result = None
   if len(lst) == 0:
	result = 0
   else:
   	avg = sum(lst)/len(lst)
   	result = int(avg)
   return result 

def rrd_ds(file):
    info = rrdtool.info(file)
    ds_re = re.compile(r'ds\[(\S+)\]')
    ds = ds_re.search(str(info))
    
    if ds:
	 result = ds.group(1)
	 return result

def rrd(file, ds):
    data = {
        'max': None,
        'avg': None
    }

    try:
        max_data = rrdtool.xport('DEF:%s=%s:%s:LAST' % (ds,file,ds), '-s', '-30d', '-e', 'now', 'XPORT:%s:last' % ds)
        start_max = max_data['meta']['start']
        end_max = max_data['meta']['end']
        ds_max = max(max_data['data'])
        ds_avg = [int(i[0]) for i in max_data['data'] if i[0] != None if int(i[0]) != 0]

        start_max_stamp = datetime.fromtimestamp(start_max)
        start_max_time = start_max_stamp.strftime('%Y-%m-%d %H:%M:%S')
        end_max_stamp = datetime.fromtimestamp(end_max)
        end_max_time = end_max_stamp.strftime('%Y-%m-%d %H:%M:%S')

        d_max = {
            'start' : None,
            'end' : None,
            'data' : None
        }

        d_avg = {
            'start' : None,
            'end' : None,
            'data' : None
        }

        d_max['start'] = start_max_time
        d_max['end'] = end_max_time
        d_max['data'] = int(ds_max[0])
        d_avg['start'] = start_max_time
        d_avg['end'] = end_max_time
        d_avg['data'] = average(ds_avg) 
        data['max'] = d_max
        data['avg'] = d_avg

	return data

    except Exception as msg:
        return data, msg

def get_contacts(filename):
    emails = None
    with open(filename, 'r') as contacts:
	emails = contacts.read().split('\n')
    return emails

def msg_template(filename):
    with open(filename, 'r') as template_file:
	template_content = template_file.read()
    return Template(template_content)

def time_log():
    time_log_stamp = datetime.now()
    log_time = time_log_stamp.strftime('%Y-%m-%d %H:%M:%S')
    return log_time

def main():
    system_dir = '/root/script/backup_rrd/'
    os.chdir(system_dir)
    host_invent = 'data/cpu_inventory.xlsx'
    data_dir = 'backup_rrd/'
    wb = load_workbook(host_invent)
    ws = wb.active

    log = open('sys.log', 'a+')
    listdir = os.listdir(data_dir)

    for n in range(3, ws.max_row+1):
    	host = ws.cell(row=n, column=1).value
    	file = None

    	for f in listdir:
            if host.lower() in f.lower() and '5min_cpu' in f.lower():
            	file = f
	    	break
            elif host.lower() in f.lower() and '6145' in f.lower():
   	    	file = f
	    	break
            elif host.lower() in f.lower() and '6013' in f.lower():
            	file = f
            	break
            else:
	    	continue
    	if file == None:
            continue
    
    	file = '%s%s' % (data_dir, file)
    	ds_info = rrd_ds(file)
    	output = rrd(file, ds_info)
    	log_output = '%s : %s --> %s\n' % (time_log(), host, output)
    	print log_output
    	log.write(log_output)
    	try:
    	    ws.cell(row=n, column=3, value=output['max']['data'])
    	    ws.cell(row=n, column=2, value=output['avg']['data'])
    	    wb.save(host_invent)
    	except:
            continue
	
    mail_account = open('mail/account.txt', 'r')
    account = mail_account.read().split(':')
    host_mail = 'smtp.gmail.com'
    port_mail = 587
    MYADD = account[0] 
    MYPASS = account[1]
    mail_msg = 'mail/message.txt'
    mail_dst = 'mail/contact_mail.txt'
    mails = get_contacts(mail_dst)
    msg_tmp = msg_template(mail_msg)
    this_month = datetime.now() - timedelta(1)
    this_month = this_month.strftime('%B') 
    print '%s : sent to --> %s' % (time_log(),mails)
    log.write('%s : sent to --> %s\n' % (time_log(), mails))
    while True:
    	try:
    	    s = smtplib.SMTP(host=host_mail, port=port_mail)
    	    s.starttls()
            break
    	except Exception as msg:
            print '%s : Connection failed : %s' % (time_log(),msg)
            log.write('%s : Connection failed : %s\n' % (time_log(), msg))
	    time.sleep(60)
    while True:
    	try:
    	    s.login(MYADD, MYPASS)
            print '%s : login successfully' % time_log()
            log.write('%s : login successfully\n' % time_log())
 	    msg = MIMEMultipart()
            msg['From'] = MYADD
	    msg['To'] = ','.join(mails)
     	    msg['Subject'] = 'CPU Report Month %s' % this_month
	    message = msg_tmp.substitute(MONTH=this_month.title())
	    msg.attach(MIMEText(message, 'plain'))
        
  	    part = MIMEBase('application', 'octet-stream')
	    part.set_payload(open('%s' % host_invent, 'rb').read())
	    encoders.encode_base64(part)
	    part.add_header('Content-Disposition', 'attachment; filename="CPU_Report.xlsx"')
	    msg.attach(part)

            s.sendmail(MYADD, mails, msg.as_string())
            print '%s : mail sent successfully' % time_log()
            log.write('%s : mail sent successfully\n' % time_log())
            del msg
            s.quit()
	    break

        except Exception as dbg:
       	   print '%s : Username or password is wrong : %s' % (time_log(), dbg)
           log.write('%s : %s\n' % (time_log(),dbg) )
	   time.sleep(5)
         
    log.close()

if __name__ == '__main__':
    main()
